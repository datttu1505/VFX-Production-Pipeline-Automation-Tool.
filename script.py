import argparse
import re
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime
import os
import subprocess
import tempfile

# Hardcoded video path
VIDEO_PATH = "FS_Winter2021_Acosta_KChaja_v2.mov"  

def filter_xytech_paths(filepath):
    pat = re.compile(r"^/hpsans\d+/.+", re.IGNORECASE)
    paths = []
    seen = set()
    
    with open(filepath) as file:
        for line in file:
            line = line.rstrip("\n")
            if pat.match(line) and line not in seen:
                paths.append(line)
                seen.add(line)
    
    return paths

def process_baselight_file(filepath):
    data = []
    with open(filepath) as file:
        for line in file:
            line = line.strip()
            if line:
                parts = line.split()
                if parts:
                    path = parts[0]
                    frames = parts[1:] if len(parts) > 1 else []
                    data.append({
                        "path": path,
                        "frames": frames
                    })
    return data

def extract_common_path(full_path):
    xytech_pattern = re.compile(r'^/hpsans\d+/production/(.+)', re.IGNORECASE)
    baselight_pattern = re.compile(r'^/baselightfilesystem\d+/(.+)', re.IGNORECASE)
    
    xytech_match = xytech_pattern.match(full_path)
    if xytech_match:
        return xytech_match.group(1)
    
    baselight_match = baselight_pattern.match(full_path)
    if baselight_match:
        return baselight_match.group(1)
    
    return None

def frame_to_timecode(frame):
    frames = frame % 24
    seconds = (frame // 24) % 60
    minutes = (frame // (24 * 60)) % 60
    hours = frame // (24 * 60 * 60)
    return f"{hours:02}:{minutes:02}:{seconds:02}:{frames:02}"

def range_to_timecode(frame_range):
    if '-' in frame_range:
        start, end = map(int, frame_range.split('-'))
        start_tc = frame_to_timecode(start)
        end_tc = frame_to_timecode(end)
        return f"{start_tc} to {end_tc}"
    else:
        frame = int(frame_range)
        return frame_to_timecode(frame)

def rangemaker(l):
    output = []
    i = 0

    while i < len(l):
        j = i
        while j < len(l) - 1 and l[j+1] == l[j] + 1:
            j += 1
        
        if j > i:
            output.append(str(l[i]) + '-' + str(l[j]))
        else:
            output.append(str(l[i]))
        
        i = j + 1

    return output

def add_padding(frame_ranges):
    padded_ranges = []
    
    for r in frame_ranges:
        if '-' in r:
            # It's a range
            start, end = map(int, r.split('-'))
            padded_start = max(1, start - 48)  # Don't go below frame 1
            padded_end = end + 48
            padded_ranges.append(f"{padded_start}-{padded_end}")
        else:
            # Single frame
            frame = int(r)
            padded_start = max(1, frame - 48)
            padded_end = frame + 48
            padded_ranges.append(f"{padded_start}-{padded_end}")
    
    return padded_ranges

def split_into_consecutive_groups(frames):
    if not frames:
        return []
    
    groups = []
    current_group = [frames[0]]
    
    for i in range(1, len(frames)):
        if frames[i] == frames[i-1] + 1:
            current_group.append(frames[i])
        else:
            groups.append(current_group)
            current_group = [frames[i]]
    
    groups.append(current_group)
    return groups

def timecode_to_seconds(timecode):
    try:
        parts = timecode.split(':')
        if len(parts) != 4:
            return None
        
        hours = int(parts[0])
        minutes = int(parts[1])
        seconds = int(parts[2])
        frames = int(parts[3])
        
        total_seconds = hours * 3600 + minutes * 60 + seconds + (frames / 24.0)
        return total_seconds
    except:
        return None
# def upload_video_to_vimeo(video_path, title=None, description=None):
#     """
#     Upload a video to Vimeo and return the video URI
#     """
#     if vimeo is None:
#         print("Error: vimeo library not installed. Run: pip install PyVimeo")
#         return None
    
#     # Vimeo credentials
#     client_id = '813b0333271adc2ec724c4b90597f5a4cecdd018'
#     client_secret = '75HvZm3of3ScVreUfE5cfDWotRLaXRZ+37aIy6rlWsGW1+cJIqlxR7mk3COQ4H519oIuwJdQUxh/dcD9Gh67QmG9TDr1Lm1gBFC7s+Ef4UsarAZ9FClEvI9pgwIcu5Dl'
#     access_token = 'f42e18071934ce5562a88ed8464a2538'
    
#     # Check if video file exists
#     if not os.path.exists(video_path):
#         print(f"Error: Video file not found: {video_path}")
#         return None
    
#     print(f"\nUploading video: {video_path}")
#     print("This may take a while depending on file size...")
    
#     try:
#         # Get file size
#         file_size = os.path.getsize(video_path)
#         print(f"File size: {file_size / (1024*1024):.2f} MB")
        
#         # Try different ways to initialize the Vimeo client
#         v = None
        
#         # Method 1: Try vimeo.VimeoClient (older library)
#         if hasattr(vimeo, 'VimeoClient'):
#             v = vimeo.VimeoClient(
#                 token=access_token,
#                 key=client_id,
#                 secret=client_secret
#             )
#         # Method 2: Try PyVimeo style
#         elif hasattr(vimeo, 'Client'):
#             v = vimeo.Client(
#                 token=access_token,
#                 key=client_id,
#                 secret=client_secret
#             )
#         else:
#             print("Error: Could not find VimeoClient or Client class")
#             print("Try: pip uninstall vimeo PyVimeo && pip install PyVimeo==1.1.0")
#             return None
        
#         # Upload the video
#         print("Starting upload...")
#         uri = v.upload(video_path, data={
#             'name': title or os.path.basename(video_path),
#             'description': description or ''
#         })
        
#         print(f"\nVideo uploaded successfully!")
#         print(f"Vimeo URI: {uri}")
        
#         # Extract video ID from URI
#         video_id = uri.split('/')[-1]
#         print(f"Video URL: https://vimeo.com/{video_id}")
        
#         return uri
    
def extract_thumbnail_ffmpeg(video_path, timecode, output_path, size="96x74"):
    """
    Extract a thumbnail from video at specific timecode using ffmpeg
    """
    try:
        # Convert timecode to seconds
        seconds = timecode_to_seconds(timecode)
        if seconds is None:
            print(f"Invalid timecode: {timecode}")
            return None
        
        # Build ffmpeg command
        cmd = [
            'ffmpeg',
            '-ss', str(seconds),  # Seek to position
            '-i', video_path,     # Input file
            '-vframes', '1',      # Extract 1 frame
            '-s', size,           # Resize to specified size
            '-y',                 # Overwrite output
            output_path
        ]
        
        # Run ffmpeg
        result = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=30
        )
        
        if result.returncode == 0 and os.path.exists(output_path):
            return output_path
        else:
            print(f"ffmpeg failed for timecode {timecode}")
            return None
            
    except subprocess.TimeoutExpired:
        print(f"ffmpeg timeout for timecode {timecode}")
        return None
    except FileNotFoundError:
        print("Error: ffmpeg not found")
        return None
    except Exception as e:
        print(f"Error extracting thumbnail: {str(e)}")
        return None

def process_video_thumbnails(db, video_path):
    if not os.path.exists(video_path):
        print(f"Warning: Video file not found: {video_path}")
        print("Skipping thumbnail generation.")
        return False
    
    print(f"\nProcessing video: {video_path}")
    print("Extracting thumbnails from timecodes...")
    
    processed_collection = db['processed']
    docs = list(processed_collection.find())
    
    if not docs:
        print("No processed data found.")
        return False
    
    # Create temporary directory for thumbnails
    thumb_dir = tempfile.mkdtemp(prefix="vfx_thumbnails_")
    print(f"Thumbnail directory: {thumb_dir}")
    
    success_count = 0
    for i, doc in enumerate(docs):
        timecode_str = doc.get('timecode', '')
        
        if not timecode_str:
            continue
        
        
        if ' to ' in timecode_str:
            first_timecode = timecode_str.split(' to ')[0].strip()
        else:
            first_timecode = timecode_str.strip()
        
        # Generate thumbnail filename
        thumb_filename = f"thumb_{i:04d}.jpg"
        thumb_path = os.path.join(thumb_dir, thumb_filename)
        
        # Extract thumbnail
        result = extract_thumbnail_ffmpeg(video_path, first_timecode, thumb_path)
        
        if result:
            # Update document with thumbnail path
            processed_collection.update_one(
                {'_id': doc['_id']},
                {'$set': {'thumbnail_path': thumb_path}}
            )
            success_count += 1
            print(f"  [{i+1}/{len(docs)}] Generated thumbnail for {first_timecode}")
        else:
            print(f"  [{i+1}/{len(docs)}] Failed to generate thumbnail for {first_timecode}")
    
    print(f"\nGenerated {success_count}/{len(docs)} thumbnails")
    return True

def export_to_excel(db, output_filename=None):
   
    processed_collection = db['processed']
    # Retrieve documents in order (sorted by first_frame during processing)
    docs = list(processed_collection.find().sort('first_frame', 1))
    
    if not docs:
        print("No processed data to export. Run --process first.")
        return
    
    # Create workbook and select active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Processed Results"
    
    # Add headers
    headers = ["Location", "Frames", "Timecode", "Thumbnail"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set row height for header
    ws.row_dimensions[1].height = 20
    
    # Add data rows with thumbnails
    row_num = 2
    for doc in docs:
        # Add data to cells
        ws.cell(row=row_num, column=1, value=doc.get('location', ''))
        ws.cell(row=row_num, column=2, value=doc.get('frames', ''))
        ws.cell(row=row_num, column=3, value=doc.get('timecode', ''))
        
        # Add thumbnail if available (Column D)
        thumb_path = doc.get('thumbnail_path')
        if thumb_path and os.path.exists(thumb_path):
            try:
                img = XLImage(thumb_path)
                # Resize image to fit cell (96x74)
                img.width = 96
                img.height = 74
                # Add image to cell D{row_num}
                ws.add_image(img, f'D{row_num}')
                # Set row height to accommodate image
                ws.row_dimensions[row_num].height = 56  # 74 pixels ≈ 56 points
            except Exception as e:
                print(f"Warning: Could not add thumbnail to row {row_num}: {str(e)}")
        
        # Apply alignment to data cells
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_num, column=3).alignment = Alignment(horizontal="center", vertical="center")
        
        row_num += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 70  # Location
    ws.column_dimensions['B'].width = 20  # Frames
    ws.column_dimensions['C'].width = 40  # Timecode
    ws.column_dimensions['D'].width = 14  # Thumbnail column
    
    # Generate filename if not provided
    if not output_filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"processed_results_{timestamp}.xlsx"
    
    # Save workbook
    wb.save(output_filename)
    print(f"\nExported {len(docs)} records to: {output_filename}")
    
    return output_filename

def process_and_match(db):

    xytech_collection = db['xytech']
    baselight_collection = db['baselight']
    processed_collection = db['processed']
    
    # Clear existing processed data
    processed_collection.delete_many({})
    
    # Get all Xytech paths
    xytech_docs = list(xytech_collection.find())
    baselight_docs = list(baselight_collection.find())
    
    # Create a mapping of common paths to baselight frames
    baselight_map = {}
    for bl_doc in baselight_docs:
        common_path = extract_common_path(bl_doc['path'])
        if common_path:
            # Convert frame strings to integers
            frames = [int(f) for f in bl_doc['frames']]
            frames.sort()
            baselight_map[common_path] = frames
    
    # Match Xytech paths with Baselight frames
    processed_docs = []
    matched_count = 0
    unmatched_count = 0
    
    for xy_doc in xytech_docs:
        common_path = extract_common_path(xy_doc['path'])
        
        if common_path and common_path in baselight_map:
            frames = baselight_map[common_path]
            
            # Split into consecutive groups
            frame_groups = split_into_consecutive_groups(frames)
            
            # Process each group separately
            for group in frame_groups:
                # Create ranges from consecutive frames
                ranges = rangemaker(group)
                
                # Add padding (48 frames before and after)
                padded_ranges = add_padding(ranges)
                
                # Convert frame ranges to timecodes
                timecodes = [range_to_timecode(r) for r in padded_ranges]
                
                # Store the first frame for sorting
                first_frame = group[0]
                
                # Create a document for each group
                processed_docs.append({
                    "location": xy_doc['path'],
                    "frames": ' '.join(padded_ranges),
                    "timecode": ' '.join(timecodes),
                    "first_frame": first_frame  # Store for sorting
                })
            
            matched_count += 1
        else:
            # Unmatched path - assign a very high frame number for sorting
            processed_docs.append({
                "location": xy_doc['path'],
                "frames": "",
                "timecode": "",
                "first_frame": float('inf')  # Sort unmatched to the end
            })
            unmatched_count += 1
    
    # Sort by first frame in ascending order
    processed_docs.sort(key=lambda x: x['first_frame'])
    
    # Insert processed documents
    if processed_docs:
        processed_collection.insert_many(processed_docs)
    
    print(f"\nProcessing Results:")
    print(f"  Matched locations: {matched_count}")
    print(f"  Unmatched locations: {unmatched_count}")
    print(f"  Total entries created: {len(processed_docs)}")
    
    return processed_docs

def main():
    parser = argparse.ArgumentParser(description='Process Baselight and Xytech files into MongoDB')
    parser.add_argument('--baselight', help='Path to Baselight file')
    parser.add_argument('--xytech', help='Path to Xytech file')
    parser.add_argument('--process', action='store_true', 
                       help='Match Xytech locations with Baselight frames and generate thumbnails')
    parser.add_argument('--export', nargs='?', const=True, metavar='FILENAME',
                       help='Export processed results to Excel (optional: specify filename)')
    parser.add_argument('--db-name', default='production_db', help='MongoDB database name')
    parser.add_argument('--mongo-uri', default='mongodb://localhost:27017/', 
                       help='MongoDB connection URI')
    
    args = parser.parse_args()
    
    # Connect to MongoDB
    client = MongoClient(args.mongo_uri)
    db = client[args.db_name]
    
    # If --export flag is set, export to Excel
    if args.export:
        filename = args.export if isinstance(args.export, str) else None
        export_to_excel(db, filename)
        client.close()
        return
    
    # If --process flag is set, match and create processed collection
    if args.process:
        print("Processing and matching Xytech with Baselight...")
        processed_docs = process_and_match(db)
        
        # Display all results
        print("\nProcessed Results:")
        print(f"{'Location':<70} {'Frames':<20} {'Timecode'}")
        print("-" * 140)
        for doc in processed_docs:
            frames_str = doc['frames'] if doc['frames'] else 'No frames'
            timecode_str = doc['timecode'] if doc['timecode'] else 'N/A'
            print(f"{doc['location']:<70} {frames_str:<20} {timecode_str}")
        
        # Process video for thumbnails using hardcoded path
        print("\nGenerating thumbnails...")
        process_video_thumbnails(db, VIDEO_PATH)
        
        client.close()
        return
    
    # Otherwise, import files if provided
    if not args.baselight or not args.xytech:
        parser.error("Either --process or both --baselight and --xytech are required")
    
    # Process and insert Xytech data (with filter)
    print("Processing Xytech file...")
    xytech_paths = filter_xytech_paths(args.xytech)
    xytech_collection = db['xytech']
    xytech_collection.delete_many({})  # Clear existing data
    
    if xytech_paths:
        xytech_docs = [{"path": path} for path in xytech_paths]
        xytech_collection.insert_many(xytech_docs)
        print(f"Inserted {len(xytech_docs)} Xytech paths")
    
    # Process and insert Baselight data
    print("Processing Baselight file...")
    baselight_data = process_baselight_file(args.baselight)
    baselight_collection = db['baselight']
    baselight_collection.delete_many({})  # Clear existing data
    
    if baselight_data:
        baselight_collection.insert_many(baselight_data)
        print(f"Inserted {len(baselight_data)} Baselight records")
    
    print("\nDone! You can now run with --process to match the data and generate thumbnails.")
    client.close()

if __name__ == "__main__":
    main()